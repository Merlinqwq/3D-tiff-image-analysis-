"""Verify merged workbooks and prove source workbooks were not modified."""

from __future__ import annotations

import csv
import hashlib
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> int:
    manifest = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    failures: list[str] = []

    for group in manifest["groups"]:
        output = Path(group["output"])
        workbook = load_workbook(output, read_only=True, data_only=False)
        if workbook.sheetnames != ["Nuclei"]:
            failures.append(f"{group['group']}: unexpected sheets {workbook.sheetnames}")
        sheet = workbook["Nuclei"]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        expected_headers = [
            "acquisition_date", "condition", "source_target_set", "source_folder",
            "image", "nucleus_id", "dapi_integrated_intensity",
            *[f"{target}_integrated_intensity" for target in group["targets"]],
        ]
        if headers != expected_headers:
            failures.append(f"{group['group']}: headers differ: {headers}")
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        if len(rows) != group["nuclei"]:
            failures.append(f"{group['group']}: expected {group['nuclei']} rows, found {len(rows)}")
        image_count = len({(row[3], row[4]) for row in rows})
        if image_count != group["images"]:
            failures.append(f"{group['group']}: expected {group['images']} images, found {image_count}")
        if any(not isinstance(value, (int, float)) for row in rows for value in row[5:]):
            failures.append(f"{group['group']}: non-numeric nucleus/intensity value")
        workbook.close()

        with Path(group["csv"]).open("r", newline="", encoding="utf-8") as handle:
            source_rows = list(csv.reader(handle))
        if len(source_rows) - 1 != len(rows):
            failures.append(f"{group['group']}: CSV/workbook row mismatch")

    changed_sources: list[str] = []
    for snapshot in manifest["source_snapshots"]:
        path = Path(snapshot["path"])
        stat = path.stat()
        if (
            stat.st_size != snapshot["size"]
            or stat.st_mtime_ns != snapshot["mtime_ns"]
            or sha256(path) != snapshot["sha256"]
        ):
            changed_sources.append(str(path))
    if changed_sources:
        failures.append(f"Modified source workbooks: {changed_sources}")

    print(f"MERGED_WORKBOOKS={len(manifest['groups'])}")
    print(f"MERGED_NUCLEI={sum(group['nuclei'] for group in manifest['groups'])}")
    print(f"SOURCE_WORKBOOKS_CHECKED={len(manifest['source_snapshots'])}")
    print(f"SOURCE_WORKBOOKS_CHANGED={len(changed_sources)}")
    print(f"FAILURES={len(failures)}")
    for failure in failures:
        print(f"FAIL: {failure}")
    return 1 if failures else 0


if __name__ == "__main__":
    raise SystemExit(main())
