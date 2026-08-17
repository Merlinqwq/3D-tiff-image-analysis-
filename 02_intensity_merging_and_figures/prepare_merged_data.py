"""Extract integrated-intensity-only tables from per-folder workbooks."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def workbook_paths(root: Path) -> list[Path]:
    return sorted(
        path
        for path in root.rglob("nuclear_intensity_results.xlsx")
        if path.parent.name == "Intensity"
    )


def acquisition_date(path: Path) -> str:
    for part in path.parts:
        match = re.search(r"(?<!\d)(\d{8})(?!\d)", part)
        if match:
            value = match.group(1)
            return f"{value[:4]}-{value[4:6]}-{value[6:]}"
    return "unknown"


def target_name_from_header(header: str) -> str:
    return header.removesuffix("_integrated_intensity")


def group_for(date: str, targets: list[str]) -> tuple[str, list[str]]:
    del date
    return "_".join(targets) if targets else "no_target", targets


def prepare(root: Path, output_dir: Path) -> Path:
    root = root.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    groups: dict[str, dict] = {}
    snapshots: list[dict] = []

    for workbook_path in workbook_paths(root):
        stat = workbook_path.stat()
        snapshots.append(
            {
                "path": str(workbook_path),
                "size": stat.st_size,
                "mtime_ns": stat.st_mtime_ns,
                "sha256": sha256(workbook_path),
            }
        )
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        nuclei_sheet = workbook["Nuclei"]
        image_sheet = workbook["Images"]
        nuclei_headers = [cell.value for cell in next(nuclei_sheet.iter_rows(max_row=1))]
        nuclei_index = {name: index for index, name in enumerate(nuclei_headers)}
        image_headers = [cell.value for cell in next(image_sheet.iter_rows(max_row=1))]
        image_index = {name: index for index, name in enumerate(image_headers)}
        image_conditions = {
            row[image_index["image"]]: row[image_index["condition"]]
            for row in image_sheet.iter_rows(min_row=2, values_only=True)
        }

        integrated_headers = [
            str(name)
            for name in nuclei_headers
            if str(name).endswith("_integrated_intensity")
            and name != "dapi_integrated_intensity"
        ]
        targets = [target_name_from_header(name) for name in integrated_headers]
        date = acquisition_date(workbook_path)
        group_name, kept_targets = group_for(date, targets)
        group = groups.setdefault(
            group_name,
            {
                "group": group_name,
                "targets": kept_targets,
                "rows": [],
                "source_workbooks": set(),
                "source_images": set(),
            },
        )
        if group["targets"] != kept_targets:
            raise ValueError(f"Inconsistent target columns for group {group_name}")

        relative_folder = str(workbook_path.parent.parent.relative_to(root))
        source_target_set = "+".join(targets)
        for values in nuclei_sheet.iter_rows(min_row=2, values_only=True):
            image_name = values[nuclei_index["image"]]
            row = {
                "acquisition_date": date,
                "condition": image_conditions[image_name],
                "source_target_set": source_target_set,
                "source_folder": relative_folder,
                "image": image_name,
                "nucleus_id": values[nuclei_index["nucleus_id"]],
                "dapi_integrated_intensity": values[nuclei_index["dapi_integrated_intensity"]],
            }
            for target in kept_targets:
                header = f"{target}_integrated_intensity"
                row[header] = values[nuclei_index[header]]
            group["rows"].append(row)
            group["source_images"].add(f"{workbook_path}|{image_name}")
        group["source_workbooks"].add(str(workbook_path))
        workbook.close()

    manifest_groups: list[dict] = []
    for group_name, group in sorted(groups.items()):
        headers = [
            "acquisition_date", "condition", "source_target_set", "source_folder",
            "image", "nucleus_id", "dapi_integrated_intensity",
            *[f"{target}_integrated_intensity" for target in group["targets"]],
        ]
        csv_path = output_dir / f"{group_name}.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            writer.writerows(group["rows"])
        manifest_groups.append(
            {
                "group": group_name,
                "targets": group["targets"],
                "csv": str(csv_path),
                "output": str(root / f"Merged_{group_name}_integrated_intensity.xlsx"),
                "nuclei": len(group["rows"]),
                "images": len(group["source_images"]),
                "source_workbooks": len(group["source_workbooks"]),
            }
        )

    manifest_path = output_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps({"groups": manifest_groups, "source_snapshots": snapshots}, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(manifest_groups, indent=2))
    return manifest_path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("root", type=Path)
    parser.add_argument("output_dir", type=Path)
    args = parser.parse_args()
    manifest = prepare(args.root, args.output_dir)
    print(f"MANIFEST={manifest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
