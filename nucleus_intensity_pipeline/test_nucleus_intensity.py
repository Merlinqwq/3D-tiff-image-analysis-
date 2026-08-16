import tempfile
import unittest
from pathlib import Path

import numpy as np
import tifffile
from openpyxl import load_workbook
from skimage import measure

from nucleus_intensity import (
    Settings,
    analyze_folder,
    infer_target_names,
    filter_labels_by_area,
    load_sum_projections,
    measure_nuclei,
    segment_nuclei,
    split_touching_nuclei,
)


class PipelineTests(unittest.TestCase):
    def test_post_watershed_area_filter_preserves_separate_labels(self):
        labels = np.zeros((120, 220), dtype=np.uint32)
        labels[10:90, 10:90] = 1  # 6,400 px: keep.
        labels[10:90, 90:170] = 2  # 6,400 px and touching label 1: keep separate.
        labels[95:105, 180:200] = 3  # 200 px: remove.
        filtered = filter_labels_by_area(labels, 5000)
        self.assertEqual(int(filtered.max()), 2)
        self.assertEqual(int(filtered[50, 50]), 1)
        self.assertEqual(int(filtered[50, 120]), 2)
        self.assertEqual(int(filtered[100, 190]), 0)

    def test_watershed_separates_touching_nuclei(self):
        yy, xx = np.ogrid[:420, :500]
        binary = np.zeros((420, 500), dtype=bool)
        for center_x, center_y in [(170, 100), (290, 100), (170, 220), (290, 220)]:
            binary |= (xx - center_x) ** 2 + (yy - center_y) ** 2 <= 70 ** 2
        binary |= (xx - 55) ** 2 + (yy - 350) ** 2 <= 55 ** 2
        self.assertEqual(int(measure.label(binary).max()), 2)
        labels = split_touching_nuclei(binary, min_distance_px=65)
        self.assertEqual(int(labels.max()), 5)

    def test_documented_segmentation_defaults(self):
        settings = Settings()
        self.assertEqual(settings.min_nucleus_area_px, 5000)
        self.assertEqual(settings.watershed_min_distance_px, 65)
        self.assertEqual(settings.min_nucleus_radius_px, 5.0)

    def test_target_name_inference(self):
        root = Path("20260730 MCF10A hypoxia normoxia")
        three_channel = root / "FBL-568 10ms NOP16-568 40ms" / "MCF10A hypoxia FBL-568 NOP61-488 DAPI Rep1" / "New folder"
        two_channel = root / "CENPB-488 50ms" / "MCF10A normoxia CENPB-488 DAPI Rep2" / "New folder"
        self.assertEqual(infer_target_names(three_channel, 3), {2: "NOP16", 3: "FBL"})
        self.assertEqual(infer_target_names(two_channel, 2), {2: "CENPB"})

    def test_fill_holes_and_remove_tiny_dapi_foci(self):
        dapi = np.zeros((64, 64), dtype=np.uint16)
        yy, xx = np.ogrid[:64, :64]
        ring = ((xx - 30) ** 2 + (yy - 30) ** 2 <= 10 ** 2) & (
            (xx - 30) ** 2 + (yy - 30) ** 2 >= 4 ** 2
        )
        dapi[ring] = 100
        dapi[5:7, 5:7] = 100  # Four-pixel random focus.
        labels, _ = segment_nuclei(
            dapi,
            Settings(min_nucleus_area_px=20, gaussian_sigma_px=0),
        )
        self.assertEqual(int(labels.max()), 1)
        self.assertGreater(int(labels[30, 30]), 0)  # Enclosed hole was filled.
        self.assertEqual(int(labels[5, 5]), 0)  # Tiny focus was removed.

    def test_one_pixel_wide_roi_boundary(self):
        from nucleus_intensity import _boundary_vertices

        labels = np.zeros((16, 32), dtype=np.uint32)
        labels[5, 3:23] = 1
        rows = _boundary_vertices(labels, "thin_object.tif")
        self.assertGreater(len(rows), 0)
        self.assertTrue(all(row["nucleus_id"] == 1 for row in rows))

    def test_zcyx_projection_and_measurement(self):
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as tmp:
            folder = (
                Path(tmp)
                / "20260816 MCF10A hypoxia normoxia"
                / "Protein568-568 Protein488-488"
                / "MCF10A hypoxia Protein568-568 Protein488-488 DAPI Rep1"
                / "New folder"
            )
            folder.mkdir(parents=True)
            data = np.zeros((4, 3, 64, 64), dtype=np.uint16)
            yy, xx = np.ogrid[:64, :64]
            nucleus_a = (xx - 18) ** 2 + (yy - 20) ** 2 <= 7 ** 2
            nucleus_b = (xx - 45) ** 2 + (yy - 42) ** 2 <= 8 ** 2
            mask = nucleus_a | nucleus_b
            data[:, 0, mask] = 100
            data[:, 1, mask] = 25
            data[:, 2, mask] = 7
            source = folder / "synthetic.tif"
            tifffile.imwrite(source, data, metadata={"axes": "ZCYX"})

            info = load_sum_projections(source)
            self.assertEqual(info.projections.shape, (3, 64, 64))
            self.assertEqual(int(info.projections[1, 20, 18]), 100)
            test_settings = Settings(min_nucleus_area_px=20, gaussian_sigma_px=0)
            labels, _ = segment_nuclei(info.projections[0], test_settings)
            rows = measure_nuclei(labels, info.projections, source.name, {2: "Protein488", 3: "Protein568"})
            self.assertEqual(len(rows), 2)
            self.assertTrue(all(abs(row["Protein488_mean_intensity"] - 100) < 1e-6 for row in rows))
            self.assertTrue(all(abs(row["Protein568_mean_intensity"] - 28) < 1e-6 for row in rows))

            workbook_path = analyze_folder(folder, test_settings)
            self.assertTrue(workbook_path.exists())
            workbook = load_workbook(workbook_path, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Nuclei", "Images", "Settings", "Errors"])
            image_headers = {cell.value: cell.column for cell in workbook["Images"][1]}
            self.assertEqual(workbook["Images"].cell(2, image_headers["nucleus_count"]).value, 2)
            nucleus_headers = [cell.value for cell in workbook["Nuclei"][1]]
            self.assertIn("Protein488_mean_intensity", nucleus_headers)
            self.assertIn("Protein568_mean_intensity", nucleus_headers)
            self.assertNotIn("ch2_mean_intensity", nucleus_headers)
            self.assertTrue((folder / "Intensity" / "Masks" / "synthetic_nuclei_label_mask.tif").exists())
            self.assertTrue((folder / "Intensity" / "ROIs" / "synthetic_roi_boundaries.csv").exists())

    def test_two_channel_czyx(self):
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as tmp:
            path = Path(tmp) / "two_channel.tiff"
            data = np.ones((2, 5, 16, 12), dtype=np.uint16)
            tifffile.imwrite(path, data, metadata={"axes": "CZYX"})
            info = load_sum_projections(path)
            self.assertEqual(info.projections.shape, (2, 16, 12))
            self.assertEqual(info.z_slices, 5)
            self.assertTrue(np.all(info.projections == 5))


if __name__ == "__main__":
    unittest.main()
