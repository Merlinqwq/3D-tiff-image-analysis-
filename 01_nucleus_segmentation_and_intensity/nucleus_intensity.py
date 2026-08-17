"""Batch nuclear segmentation and per-nucleus fluorescence quantification.

The first TIFF channel is treated as DAPI. Each channel is sum-projected over
Z, nuclei are segmented from the DAPI projection with Otsu thresholding, and
channel 2/3 intensities are measured within every nuclear label.
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
import traceback
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable

import numpy as np
import pandas as pd
import tifffile
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import ndimage as ndi
from skimage import feature, filters, measure, morphology, segmentation


TIFF_SUFFIXES = {".tif", ".tiff"}
OUTPUT_DIRECTORY_NAMES = {
    "intensity", "projections", "masks", "rois", "channel_analysis_results",
    "condensate_thresholding", "colocalization_analysis",
}
LOG = logging.getLogger("nucleus_intensity")


@dataclass(frozen=True)
class Settings:
    segmentation_channel: int = 1
    min_nucleus_area_px: int = 5000
    gaussian_sigma_px: float = 1.0
    fill_holes: bool = True
    clear_border: bool = False
    split_touching_nuclei: bool = True
    watershed_min_distance_px: int = 65
    min_nucleus_radius_px: float = 5.0


@dataclass(frozen=True)
class ProjectionInfo:
    projections: np.ndarray  # C, Y, X
    source_axes: str
    source_shape: tuple[int, ...]
    channel_count: int
    z_slices: int


TARGET_PATTERN = re.compile(r"(?<![A-Za-z0-9])([A-Za-z][A-Za-z0-9]*)\s*-?\s*(488|568)(?!\d)", re.IGNORECASE)
TARGET_NAME_CORRECTIONS: dict[str, str] = {}


def _target_pairs(text: str) -> list[tuple[str, int]]:
    """Extract ordered (protein, wavelength) pairs from a folder name."""
    pairs: list[tuple[str, int]] = []
    seen: set[tuple[str, int]] = set()
    for match in TARGET_PATTERN.finditer(text):
        raw_name = match.group(1)
        corrected_name = TARGET_NAME_CORRECTIONS.get(raw_name.casefold(), raw_name)
        pair = (corrected_name, int(match.group(2)))
        normalized = (pair[0].casefold(), pair[1])
        if normalized not in seen and pair[0].casefold() != "dapi":
            pairs.append(pair)
            seen.add(normalized)
    return pairs


def infer_target_names(input_folder: Path, channel_count: int) -> dict[int, str]:
    """Infer channel labels from the condition folder, then its marker folder.

    For 3-channel data, 488 maps to channel 2 and 568 maps to channel 3.
    For 2-channel data, the sole non-DAPI target maps to channel 2 regardless
    of the wavelength written in the folder name.
    """
    source_names = [input_folder.parent.name, input_folder.parent.parent.name]
    candidates = [_target_pairs(name) for name in source_names]

    if channel_count == 2:
        for pairs in candidates:
            unique_names = list(dict.fromkeys(name for name, _ in pairs))
            if len(unique_names) == 1:
                return {2: unique_names[0]}
        return {2: "ch2"}

    for pairs in candidates:
        by_wavelength = {wavelength: name for name, wavelength in pairs}
        if 488 in by_wavelength and 568 in by_wavelength:
            return {2: by_wavelength[488], 3: by_wavelength[568]}
    return {2: "ch2", 3: "ch3"}


def _target_column_prefix(target_name: str) -> str:
    prefix = re.sub(r"[^A-Za-z0-9]+", "_", target_name).strip("_")
    return prefix or "target"


def infer_condition(path: Path) -> str:
    """Infer a condition label from the closest matching ancestor."""
    for part in reversed(path.parts):
        text = part.casefold()
        if "normoxia" in text:
            return "normoxia"
        if "hypoxia" in text:
            return "hypoxia"
    return "unspecified"


def _normalize_axes(array: np.ndarray, axes: str) -> tuple[np.ndarray, list[str]]:
    """Remove singleton non-spatial axes and infer a missing channel axis."""
    axis_names = list(axes.upper())
    if len(axis_names) != array.ndim:
        raise ValueError(f"TIFF axes {axes!r} do not match shape {array.shape}.")

    for index in reversed(range(array.ndim)):
        if array.shape[index] == 1 and axis_names[index] not in {"Y", "X"}:
            array = np.squeeze(array, axis=index)
            axis_names.pop(index)

    if "Y" not in axis_names or "X" not in axis_names:
        raise ValueError(f"Could not identify Y and X axes from TIFF axes {axes!r}.")

    if "C" not in axis_names:
        candidates = [
            i
            for i, (name, size) in enumerate(zip(axis_names, array.shape))
            if name not in {"Y", "X"} and size in (2, 3)
        ]
        if len(candidates) != 1:
            raise ValueError(
                "Could not identify the channel axis. Export TIFF with C in its "
                f"axes metadata; found axes={''.join(axis_names)}, shape={array.shape}."
            )
        axis_names[candidates[0]] = "C"

    return array, axis_names


def load_sum_projections(path: Path) -> ProjectionInfo:
    """Read the first TIFF series and return C,Y,X sum projections."""
    raw_memmap: np.memmap | None = None
    with tifffile.TiffFile(path) as tif:
        series = tif.series[0]
        source_axes = series.axes
        source_shape = tuple(int(v) for v in series.shape)
        # A real microscopy stack can be several GB. A tifffile-managed memmap
        # avoids holding the complete Z stack in RAM before projection.
        array = series.asarray(out="memmap")
        if isinstance(array, np.memmap):
            raw_memmap = array

    try:
        array, axes = _normalize_axes(array, source_axes)
        c_axis = axes.index("C")
        channel_count = int(array.shape[c_axis])
        if channel_count not in (2, 3):
            raise ValueError(f"Expected 2 or 3 channels, found {channel_count}.")

        z_slices = int(array.shape[axes.index("Z")]) if "Z" in axes else 1
        projection_axes = [i for i, name in enumerate(axes) if name not in {"C", "Y", "X"}]
        projected = (
            array.sum(axis=tuple(projection_axes), dtype=np.uint64)
            if projection_axes
            else np.asarray(array, dtype=np.uint64)
        )
        kept_axes = [name for name in axes if name in {"C", "Y", "X"}]
        projected = np.moveaxis(
            projected,
            [kept_axes.index("C"), kept_axes.index("Y"), kept_axes.index("X")],
            [0, 1, 2],
        )
        if projected.max(initial=0) <= np.iinfo(np.uint32).max:
            projected = projected.astype(np.uint32)
    finally:
        if raw_memmap is not None:
            temporary_path = Path(raw_memmap.filename)
            raw_memmap._mmap.close()
            if temporary_path.resolve() != path.resolve():
                temporary_path.unlink(missing_ok=True)

    return ProjectionInfo(projected, source_axes, source_shape, channel_count, z_slices)


def split_touching_nuclei(
    binary: np.ndarray,
    min_distance_px: int,
    min_nucleus_radius_px: float = 5.0,
) -> np.ndarray:
    """Separate touching nuclei with marker-controlled distance watershed."""
    connected = measure.label(binary, connectivity=2)
    if connected.max(initial=0) == 0:
        return connected.astype(np.uint32)
    distance = ndi.distance_transform_edt(binary)
    coordinates = feature.peak_local_max(
        distance,
        labels=connected,
        min_distance=min_distance_px,
        threshold_abs=min_nucleus_radius_px,
        exclude_border=False,
    )
    marker_mask = np.zeros(binary.shape, dtype=bool)
    marker_mask[tuple(coordinates.T)] = True
    markers = measure.label(marker_mask, connectivity=2)
    labels = segmentation.watershed(-distance, markers, mask=binary)
    return labels.astype(np.uint32)


def filter_labels_by_area(labels: np.ndarray, min_area_px: int) -> np.ndarray:
    """Remove small labeled regions and renumber survivors without merging them."""
    counts = np.bincount(labels.ravel())
    keep = np.flatnonzero(counts >= min_area_px)
    keep = keep[keep != 0]
    mapping = np.zeros(len(counts), dtype=np.uint32)
    mapping[keep] = np.arange(1, len(keep) + 1, dtype=np.uint32)
    return mapping[labels]


def segment_nuclei(dapi_projection: np.ndarray, settings: Settings) -> tuple[np.ndarray, float]:
    """Segment DAPI-positive nuclei with global Otsu and connected components."""
    image = np.asarray(dapi_projection, dtype=np.float64)
    finite = image[np.isfinite(image)]
    if finite.size == 0 or np.ptp(finite) == 0:
        raise ValueError("DAPI projection has no usable intensity variation.")

    smoothed = filters.gaussian(image, sigma=settings.gaussian_sigma_px, preserve_range=True)
    threshold = float(filters.threshold_otsu(smoothed[np.isfinite(smoothed)]))
    binary = smoothed > threshold
    # Match the microscopy "Fill Holes" operation before extracting ROIs.
    if settings.fill_holes:
        binary = ndi.binary_fill_holes(binary)
    # Remove isolated DAPI foci/noise using a user-configurable pixel-area filter.
    binary = morphology.remove_small_objects(binary, min_size=settings.min_nucleus_area_px)
    binary = morphology.binary_closing(binary, morphology.disk(1))
    if settings.clear_border:
        binary = segmentation.clear_border(binary)

    labels = (
        split_touching_nuclei(
            binary,
            settings.watershed_min_distance_px,
            settings.min_nucleus_radius_px,
        )
        if settings.split_touching_nuclei
        else measure.label(binary, connectivity=2).astype(np.uint32)
    )
    labels = filter_labels_by_area(labels, settings.min_nucleus_area_px)
    return labels, threshold


def _boundary_vertices(labels: np.ndarray, image_name: str) -> list[dict]:
    rows: list[dict] = []
    for prop in measure.regionprops(labels):
        row_offset, col_offset = prop.bbox[0], prop.bbox[1]
        # Padding supports valid but thin 1xN/Nx1 objects and closes contours
        # for objects that touch the edge of their local bounding box.
        padded_mask = np.pad(prop.image, 1, mode="constant", constant_values=False)
        for boundary_id, contour in enumerate(measure.find_contours(padded_mask, 0.5), start=1):
            for vertex_id, (row, col) in enumerate(contour, start=1):
                rows.append(
                    {
                        "image": image_name,
                        "nucleus_id": int(prop.label),
                        "boundary_id": boundary_id,
                        "vertex_id": vertex_id,
                        "x_px": float(col - 1 + col_offset),
                        "y_px": float(row - 1 + row_offset),
                    }
                )
    return rows


def measure_nuclei(
    labels: np.ndarray,
    projections: np.ndarray,
    image_name: str,
    target_names: dict[int, str] | None = None,
    segmentation_channel: int = 1,
) -> list[dict]:
    """Return one flat, machine-readable record per nucleus."""
    if not 1 <= segmentation_channel <= projections.shape[0]:
        raise ValueError(
            f"Segmentation channel {segmentation_channel} is outside 1..{projections.shape[0]}"
        )
    dapi = projections[segmentation_channel - 1]
    props = measure.regionprops(labels, intensity_image=dapi)
    rows: list[dict] = []
    for prop in props:
        nucleus_mask = prop.image
        nucleus_slice = prop.slice
        min_row, min_col, max_row, max_col = prop.bbox
        row = {
            "image": image_name,
            "nucleus_id": int(prop.label),
            "area_px": int(prop.area),
            "centroid_x_px": float(prop.centroid[1]),
            "centroid_y_px": float(prop.centroid[0]),
            "bbox_x_min_px": int(min_col),
            "bbox_y_min_px": int(min_row),
            "bbox_x_max_exclusive_px": int(max_col),
            "bbox_y_max_exclusive_px": int(max_row),
            "dapi_mean_intensity": float(np.mean(dapi[nucleus_slice][nucleus_mask])),
            "dapi_integrated_intensity": float(np.sum(dapi[nucleus_slice][nucleus_mask], dtype=np.float64)),
        }
        for channel_index in range(projections.shape[0]):
            if channel_index == segmentation_channel - 1:
                continue
            values = projections[channel_index][nucleus_slice][nucleus_mask]
            channel_number = channel_index + 1
            target = (target_names or {}).get(channel_number, f"ch{channel_number}")
            prefix = _target_column_prefix(target)
            row[f"{prefix}_mean_intensity"] = float(np.mean(values))
            row[f"{prefix}_integrated_intensity"] = float(np.sum(values, dtype=np.float64))
            row[f"{prefix}_median_intensity"] = float(np.median(values))
            row[f"{prefix}_min_intensity"] = float(np.min(values))
            row[f"{prefix}_max_intensity"] = float(np.max(values))
        rows.append(row)
    return rows


def _save_outputs(
    output_dir: Path,
    stem: str,
    projections: np.ndarray,
    labels: np.ndarray,
    roi_rows: list[dict],
) -> dict[str, str]:
    projection_dir = output_dir / "Projections"
    mask_dir = output_dir / "Masks"
    roi_dir = output_dir / "ROIs"
    projection_dir.mkdir(parents=True, exist_ok=True)
    mask_dir.mkdir(parents=True, exist_ok=True)
    roi_dir.mkdir(parents=True, exist_ok=True)

    projection_path = projection_dir / f"{stem}_sum_projection.tif"
    binary_path = mask_dir / f"{stem}_nuclei_binary_mask.tif"
    labels_path = mask_dir / f"{stem}_nuclei_label_mask.tif"
    roi_path = roi_dir / f"{stem}_roi_boundaries.csv"

    tifffile.imwrite(projection_path, projections, metadata={"axes": "CYX"}, photometric="minisblack")
    tifffile.imwrite(binary_path, (labels > 0).astype(np.uint8) * 255, photometric="minisblack")
    tifffile.imwrite(labels_path, labels, photometric="minisblack")
    pd.DataFrame(
        roi_rows,
        columns=["image", "nucleus_id", "boundary_id", "vertex_id", "x_px", "y_px"],
    ).to_csv(roi_path, index=False)

    return {
        "projection_file": str(projection_path.relative_to(output_dir)),
        "binary_mask_file": str(binary_path.relative_to(output_dir)),
        "label_mask_file": str(labels_path.relative_to(output_dir)),
        "roi_boundary_file": str(roi_path.relative_to(output_dir)),
    }


def _format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:200]]
            width = min(max(max((len(v) for v in values), default=0) + 2, 11), 42)
            header = str(column_cells[0].value or "")
            if header == "image":
                width = 55
            elif header.endswith("_file") or header == "folder":
                width = 60
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
            if header.endswith(("_mean_intensity", "_median_intensity")) or header in {
                "centroid_x_px", "centroid_y_px", "otsu_threshold"
            }:
                for cell in column_cells[1:]:
                    cell.number_format = "#,##0.00"
            elif header.endswith(("_integrated_intensity", "_min_intensity", "_max_intensity")) or header in {
                "nucleus_id", "area_px", "bbox_x_min_px", "bbox_y_min_px",
                "bbox_x_max_exclusive_px", "bbox_y_max_exclusive_px", "channel_count",
                "z_slices", "nucleus_count",
            }:
                for cell in column_cells[1:]:
                    cell.number_format = "#,##0"
    workbook.save(path)


def write_excel(
    output_dir: Path,
    nucleus_rows: list[dict],
    image_rows: list[dict],
    settings: Settings,
    errors: list[dict],
) -> Path:
    workbook_path = output_dir / "nuclear_intensity_results.xlsx"
    base_nuclei_columns = [
        "image", "nucleus_id", "area_px", "centroid_x_px", "centroid_y_px",
        "bbox_x_min_px", "bbox_y_min_px", "bbox_x_max_exclusive_px", "bbox_y_max_exclusive_px",
        "dapi_mean_intensity", "dapi_integrated_intensity",
    ]
    measured_columns = list(dict.fromkeys(key for row in nucleus_rows for key in row))
    nuclei_columns = base_nuclei_columns + [key for key in measured_columns if key not in base_nuclei_columns]
    image_columns = [
        "image", "condition", "segmentation_channel", "channel_2_target", "channel_3_target",
        "source_axes", "source_shape",
        "channel_count", "z_slices", "otsu_threshold",
        "nucleus_count", "projection_file", "binary_mask_file", "label_mask_file", "roi_boundary_file",
    ]
    settings_rows = [
        {"parameter": key, "value": value, "description": description}
        for (key, value), description in zip(
            asdict(settings).items(),
            [
                "One-based channel used to segment nuclei or other primary objects.",
                "Remove connected DAPI objects smaller than this area.",
                "Gaussian smoothing before global Otsu thresholding.",
                "Fill internal holes in the thresholded segmentation mask.",
                "Exclude nuclei touching an image border.",
                "Separate touching DAPI nuclei with marker-controlled watershed.",
                "Minimum distance between watershed nuclear-center markers.",
                "Reject thresholded objects without this minimum interior radius.",
            ],
        )
    ]
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(nucleus_rows).reindex(columns=nuclei_columns).to_excel(writer, sheet_name="Nuclei", index=False)
        pd.DataFrame(image_rows).reindex(columns=image_columns).to_excel(writer, sheet_name="Images", index=False)
        pd.DataFrame(settings_rows).to_excel(writer, sheet_name="Settings", index=False)
        pd.DataFrame(errors, columns=["image", "error"]).to_excel(writer, sheet_name="Errors", index=False)
    _format_workbook(workbook_path)
    return workbook_path


def analyze_folder(
    input_folder: Path,
    settings: Settings | None = None,
    progress: Callable[[str], None] | None = None,
    channel_names: dict[int, str] | None = None,
) -> Path:
    settings = settings or Settings()
    input_folder = Path(input_folder).resolve()
    output_dir = input_folder / "Intensity"
    output_dir.mkdir(exist_ok=True)
    tiff_files = sorted(
        path for path in input_folder.iterdir()
        if path.is_file() and path.suffix.lower() in TIFF_SUFFIXES
    )
    if not tiff_files:
        raise FileNotFoundError(f"No .tif or .tiff files found directly in {input_folder}")

    nucleus_rows: list[dict] = []
    image_rows: list[dict] = []
    errors: list[dict] = []
    for index, path in enumerate(tiff_files, start=1):
        message = f"[{index}/{len(tiff_files)}] {path.name}"
        LOG.info(message)
        if progress:
            progress(message)
        try:
            info = load_sum_projections(path)
            if not 1 <= settings.segmentation_channel <= info.channel_count:
                raise ValueError(
                    f"--segmentation-channel must be within 1..{info.channel_count} for {path.name}"
                )
            target_names = infer_target_names(input_folder, info.channel_count)
            target_names.update(channel_names or {})
            labels, threshold = segment_nuclei(
                info.projections[settings.segmentation_channel - 1], settings
            )
            measurements = measure_nuclei(
                labels, info.projections, path.name, target_names,
                settings.segmentation_channel,
            )
            rois = _boundary_vertices(labels, path.name)
            output_files = _save_outputs(output_dir, path.stem, info.projections, labels, rois)
            nucleus_rows.extend(measurements)
            image_rows.append(
                {
                    "image": path.name,
                    "condition": infer_condition(input_folder),
                    "segmentation_channel": settings.segmentation_channel,
                    "channel_2_target": target_names.get(2),
                    "channel_3_target": target_names.get(3),
                    "source_axes": info.source_axes,
                    "source_shape": json.dumps(info.source_shape),
                    "channel_count": info.channel_count,
                    "z_slices": info.z_slices,
                    "otsu_threshold": threshold,
                    "nucleus_count": len(measurements),
                    **output_files,
                }
            )
        except Exception as exc:  # Continue the batch and report the failed image.
            LOG.exception("Failed to process %s", path)
            errors.append({"image": path.name, "error": f"{type(exc).__name__}: {exc}"})

    workbook = write_excel(output_dir, nucleus_rows, image_rows, settings, errors)
    if not image_rows:
        raise RuntimeError(f"All {len(tiff_files)} TIFF files failed. See {workbook} Errors sheet.")
    return workbook


def find_analysis_folders(batch_root: Path, folder_name: str | None = None) -> list[Path]:
    """Find leaf-most directories containing TIFF files, regardless of folder name.

    ``folder_name`` can optionally restrict recursive discovery to a chosen
    directory name. Known output directories are always ignored.
    """
    root = Path(batch_root).resolve()
    candidates = [root, *(path for path in root.rglob("*") if path.is_dir())]
    direct_tiff_folders: list[Path] = []
    for path in candidates:
        try:
            relative_parts = path.relative_to(root).parts
        except ValueError:
            continue
        if any(part.casefold() in OUTPUT_DIRECTORY_NAMES for part in relative_parts):
            continue
        if folder_name and path.name.casefold() != folder_name.casefold():
            continue
        if any(
            item.is_file() and item.suffix.casefold() in TIFF_SUFFIXES
            for item in path.iterdir()
        ):
            direct_tiff_folders.append(path.resolve())
    if folder_name:
        return sorted(set(direct_tiff_folders))
    # If both a parent and its descendant contain TIFFs, default to the deeper
    # folder to avoid processing duplicated exports in nested layouts.
    return sorted(
        path for path in set(direct_tiff_folders)
        if not any(path != other and path in other.parents for other in direct_tiff_folders)
    )


def _source_snapshot(batch_root: Path) -> dict[str, tuple[int, int]]:
    return {
        str(path.resolve()): (path.stat().st_size, path.stat().st_mtime_ns)
        for folder in find_analysis_folders(batch_root)
        for path in folder.iterdir()
        if path.is_file() and path.suffix.casefold() in TIFF_SUFFIXES
    }


def write_batch_count_qa(
    batch_root: Path,
    summary: list[dict],
    expected_count: int | None = None,
    exempt_dates: set[str] | None = None,
) -> Path:
    """Write target-by-treatment nucleus-count QA without changing any ROIs."""
    exempt_dates = exempt_dates or set()
    records: list[dict] = []
    for item in summary:
        if item["status"] != "complete":
            continue
        workbook_path = Path(item["workbook"])
        images = pd.read_excel(workbook_path, sheet_name="Images")
        date_match = re.search(r"(?<!\d)(\d{8})(?!\d)", str(workbook_path))
        date = date_match.group(1) if date_match else "unknown"
        for row in images.itertuples(index=False):
            targets = [
                str(value)
                for value in (row.channel_2_target, row.channel_3_target)
                if pd.notna(value) and str(value).strip()
            ]
            records.append(
                {
                    "date": date,
                    "target_group": "+".join(targets),
                    "condition": row.condition,
                    "nucleus_count": int(row.nucleus_count),
                }
            )
    frame = pd.DataFrame(records)
    if frame.empty:
        qa = pd.DataFrame(
            columns=["date", "target_group", "condition", "nucleus_count", "expected_count", "qa_status"]
        )
    else:
        qa = (
            frame.groupby(["date", "target_group", "condition"], as_index=False)["nucleus_count"]
            .sum()
            .sort_values(["target_group", "date", "condition"])
        )
        qa["expected_count"] = expected_count
        if expected_count is None:
            qa["qa_status"] = "NOT_CONFIGURED"
        else:
            qa["qa_status"] = np.where(
                qa["date"].isin(exempt_dates),
                "EXEMPT_DATE",
                np.where(qa["nucleus_count"] == expected_count, "PASS", "REVIEW"),
            )
    output_path = Path(batch_root).resolve() / "nuclear_intensity_count_QA.csv"
    qa.to_csv(output_path, index=False)
    return output_path


def analyze_batch(
    batch_root: Path,
    settings: Settings | None = None,
    progress: Callable[[str], None] | None = None,
    excluded_targets: set[str] | None = None,
    folder_name: str | None = None,
    channel_names: dict[int, str] | None = None,
    expected_count: int | None = None,
    qa_exempt_dates: set[str] | None = None,
) -> list[dict]:
    settings = settings or Settings()
    folders = find_analysis_folders(batch_root, folder_name)
    if not folders:
        suffix = f" named {folder_name!r}" if folder_name else ""
        raise FileNotFoundError(f"No TIFF-containing input folder{suffix} found under {batch_root}")
    source_before = _source_snapshot(batch_root)
    summary: list[dict] = []
    exclusions = {name.casefold() for name in (excluded_targets or set())}
    for index, folder in enumerate(folders, start=1):
        message = f"DATASET [{index}/{len(folders)}] {folder}"
        LOG.info(message)
        if progress:
            progress(message)
        if exclusions and any(
            excluded in folder.as_posix().casefold()
            for excluded in exclusions
        ):
            summary.append(
                {"folder": str(folder), "status": "excluded", "workbook": "", "error": "User-excluded target"}
            )
            continue
        try:
            workbook = analyze_folder(folder, settings, progress, channel_names)
            summary.append({"folder": str(folder), "status": "complete", "workbook": str(workbook), "error": ""})
        except Exception as exc:
            LOG.exception("Dataset failed: %s", folder)
            summary.append(
                {"folder": str(folder), "status": "failed", "workbook": "", "error": f"{type(exc).__name__}: {exc}"}
            )
    summary_path = Path(batch_root).resolve() / "nuclear_intensity_batch_summary.csv"
    pd.DataFrame(summary).to_csv(summary_path, index=False)
    qa_path = write_batch_count_qa(batch_root, summary, expected_count, qa_exempt_dates)
    qa_frame = pd.read_csv(qa_path)
    review_count = int((qa_frame["qa_status"] == "REVIEW").sum()) if not qa_frame.empty else 0
    if review_count:
        LOG.warning("Nucleus-count QA flagged %d non-exempt target/treatment groups: %s", review_count, qa_path)
    else:
        LOG.info("Nucleus-count QA passed for all non-exempt target/treatment groups: %s", qa_path)
    source_after = _source_snapshot(batch_root)
    integrity_rows = []
    for source_path in sorted(set(source_before) | set(source_after)):
        before = source_before.get(source_path)
        after = source_after.get(source_path)
        status = "unchanged" if before == after else "changed_or_missing"
        integrity_rows.append(
            {
                "source_tiff": source_path,
                "size_bytes_before": before[0] if before else None,
                "size_bytes_after": after[0] if after else None,
                "mtime_ns_before": before[1] if before else None,
                "mtime_ns_after": after[1] if after else None,
                "status": status,
            }
        )
    pd.DataFrame(integrity_rows).to_csv(
        Path(batch_root).resolve() / "nuclear_intensity_source_integrity.csv", index=False
    )
    return summary


def choose_folder_gui() -> Path | None:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    selected = filedialog.askdirectory(title="Choose folder containing 2- or 3-channel TIFF images")
    root.destroy()
    return Path(selected) if selected else None


def choose_min_area_gui(default: int) -> int | None:
    import tkinter as tk
    from tkinter import simpledialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    value = simpledialog.askinteger(
        "DAPI size filter",
        "Minimum nucleus area in pixels\n(smaller DAPI foci will be removed):",
        initialvalue=default,
        minvalue=1,
        parent=root,
    )
    root.destroy()
    return value


def _show_message(title: str, message: str, error: bool = False) -> None:
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        (messagebox.showerror if error else messagebox.showinfo)(title, message)
        root.destroy()
    except Exception:
        print(f"{title}: {message}", file=sys.stderr if error else sys.stdout)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-folder", type=Path, help="Skip the folder picker and analyze this folder.")
    parser.add_argument(
        "--batch-root", type=Path,
        help="Recursively process leaf-most TIFF-containing folders, regardless of folder name.",
    )
    parser.add_argument(
        "--folder-name",
        help="Optional batch restriction: process only TIFF folders with this exact directory name.",
    )
    parser.add_argument(
        "--exclude-target",
        action="append",
        default=[],
        help="Skip batch folders whose path contains this target name; may be repeated.",
    )
    parser.add_argument(
        "--segmentation-channel", type=int, default=Settings.segmentation_channel,
        help="One-based channel used for object segmentation (default: 1).",
    )
    parser.add_argument(
        "--channel-name", action="append", default=[], metavar="CHANNEL=NAME",
        help="Override an output channel name, for example --channel-name 2=ProteinA. Repeatable.",
    )
    parser.add_argument("--min-area", type=int, default=Settings.min_nucleus_area_px)
    parser.add_argument("--sigma", type=float, default=Settings.gaussian_sigma_px)
    parser.add_argument("--no-fill-holes", action="store_true")
    parser.add_argument("--clear-border", action="store_true")
    parser.add_argument("--no-watershed", action="store_true", help="Disable separation of touching nuclei.")
    parser.add_argument(
        "--watershed-min-distance",
        type=int,
        default=Settings.watershed_min_distance_px,
        help="Minimum center-to-center marker distance in pixels (default: 65).",
    )
    parser.add_argument(
        "--min-nucleus-radius",
        type=float,
        default=Settings.min_nucleus_radius_px,
        help="Minimum interior distance in pixels for a nuclear marker (default: 5).",
    )
    parser.add_argument(
        "--expected-count", type=int,
        help="Optional expected total object count per target/condition for QA only.",
    )
    parser.add_argument(
        "--qa-exempt-date", action="append", default=[],
        help="Date label exempt from expected-count QA; repeatable.",
    )
    return parser.parse_args()


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    args = parse_args()
    channel_names: dict[int, str] = {}
    for item in args.channel_name:
        try:
            channel_text, name = item.split("=", 1)
            channel = int(channel_text)
        except ValueError as exc:
            raise SystemExit(f"Invalid --channel-name {item!r}; use CHANNEL=NAME") from exc
        if channel < 1 or not name.strip():
            raise SystemExit(f"Invalid --channel-name {item!r}; use a positive channel and non-empty name")
        channel_names[channel] = name.strip()
    if args.batch_root and args.input_folder:
        raise SystemExit("Use either --batch-root or --input-folder, not both.")
    if args.batch_root:
        settings = Settings(
            segmentation_channel=args.segmentation_channel,
            min_nucleus_area_px=args.min_area,
            gaussian_sigma_px=args.sigma,
            fill_holes=not args.no_fill_holes,
            clear_border=args.clear_border,
            split_touching_nuclei=not args.no_watershed,
            watershed_min_distance_px=args.watershed_min_distance,
            min_nucleus_radius_px=args.min_nucleus_radius,
        )
        summary = analyze_batch(
            args.batch_root, settings, excluded_targets=set(args.exclude_target),
            folder_name=args.folder_name, channel_names=channel_names,
            expected_count=args.expected_count,
            qa_exempt_dates=set(args.qa_exempt_date),
        )
        completed = sum(row["status"] == "complete" for row in summary)
        failed = sum(row["status"] == "failed" for row in summary)
        excluded = sum(row["status"] == "excluded" for row in summary)
        LOG.info("Batch complete: %d datasets complete, %d failed, %d excluded", completed, failed, excluded)
        return 1 if failed else 0
    using_gui = args.input_folder is None
    folder = args.input_folder or choose_folder_gui()
    if folder is None:
        return 0
    min_area = choose_min_area_gui(args.min_area) if using_gui else args.min_area
    if min_area is None:
        return 0
    settings = Settings(
        segmentation_channel=args.segmentation_channel,
        min_nucleus_area_px=min_area,
        gaussian_sigma_px=args.sigma,
        fill_holes=not args.no_fill_holes,
        clear_border=args.clear_border,
        split_touching_nuclei=not args.no_watershed,
        watershed_min_distance_px=args.watershed_min_distance,
        min_nucleus_radius_px=args.min_nucleus_radius,
    )
    try:
        workbook = analyze_folder(folder, settings, channel_names=channel_names)
    except Exception as exc:
        LOG.debug(traceback.format_exc())
        _show_message("Nuclear intensity analysis failed", str(exc), error=True)
        return 1
    if using_gui:
        _show_message("Analysis complete", f"Results saved to:\n{workbook}")
    else:
        print(f"Analysis complete: {workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
