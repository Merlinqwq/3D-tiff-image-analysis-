"""Audit completed nuclear-intensity batch outputs without reading image pixels."""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
import tifffile
from openpyxl import load_workbook

from nucleus_intensity import TIFF_SUFFIXES, find_analysis_folders, infer_target_names


def audit(batch_root: Path, excluded_targets: set[str]) -> tuple[Path, list[dict]]:
    root = batch_root.resolve()
    exclusions = {value.casefold() for value in excluded_targets}
    rows: list[dict] = []

    for folder in find_analysis_folders(root):
        tiffs = sorted(
            item for item in folder.iterdir()
            if item.is_file() and item.suffix.casefold() in TIFF_SUFFIXES
        )
        excluded = any(value in folder.as_posix().casefold() for value in exclusions)
        if excluded:
            rows.append(
                {
                    "folder": str(folder),
                    "status": "excluded",
                    "source_tiffs": len(tiffs),
                    "workbook_images": None,
                    "workbook_errors": None,
                    "protein_headers_ok": None,
                    "output_counts_ok": None,
                    "notes": "User-excluded target",
                }
            )
            continue

        workbook_path = folder / "Intensity" / "nuclear_intensity_results.xlsx"
        problems: list[str] = []
        workbook_images = 0
        workbook_errors = 0
        protein_headers_ok = False
        output_counts_ok = False
        if not workbook_path.exists():
            problems.append("missing workbook")
        else:
            workbook = load_workbook(workbook_path, read_only=True, data_only=False)
            workbook_images = max(0, workbook["Images"].max_row - 1)
            workbook_errors = max(0, workbook["Errors"].max_row - 1)
            nuclei_headers = [cell.value for cell in next(workbook["Nuclei"].iter_rows(max_row=1))]
            image_headers = [cell.value for cell in next(workbook["Images"].iter_rows(max_row=1))]
            image_header_index = {name: index for index, name in enumerate(image_headers)}
            recorded_images = {
                row[image_header_index["image"]].value
                for row in workbook["Images"].iter_rows(min_row=2)
            }
            if recorded_images != {path.name for path in tiffs}:
                problems.append("workbook/source image list mismatch")

            if tiffs:
                with tifffile.TiffFile(tiffs[0]) as tif:
                    series = tif.series[0]
                    channel_count = int(series.shape[series.axes.index("C")])
                targets = infer_target_names(folder, channel_count)
                expected_headers = {
                    f"{target}_{stat}_intensity"
                    for target in targets.values()
                    for stat in ("mean", "integrated", "median", "min", "max")
                }
                protein_headers_ok = expected_headers.issubset(nuclei_headers) and not any(
                    str(header).startswith(("ch2_", "ch3_")) for header in nuclei_headers
                )
                if not protein_headers_ok:
                    problems.append("protein measurement headers incorrect")
            workbook.close()

            output_dir = folder / "Intensity"
            projection_count = len(list((output_dir / "Projections").glob("*_sum_projection.tif")))
            binary_count = len(list((output_dir / "Masks").glob("*_nuclei_binary_mask.tif")))
            label_count = len(list((output_dir / "Masks").glob("*_nuclei_label_mask.tif")))
            roi_count = len(list((output_dir / "ROIs").glob("*_roi_boundaries.csv")))
            output_counts_ok = all(
                count == len(tiffs)
                for count in (projection_count, binary_count, label_count, roi_count)
            )
            if not output_counts_ok:
                problems.append("projection/mask/ROI output count mismatch")

        if workbook_images != len(tiffs):
            problems.append("workbook image count mismatch")
        if workbook_errors:
            problems.append(f"{workbook_errors} image errors")
        rows.append(
            {
                "folder": str(folder),
                "status": "pass" if not problems else "fail",
                "source_tiffs": len(tiffs),
                "workbook_images": workbook_images,
                "workbook_errors": workbook_errors,
                "protein_headers_ok": protein_headers_ok,
                "output_counts_ok": output_counts_ok,
                "notes": "; ".join(dict.fromkeys(problems)),
            }
        )

    output_path = root / "nuclear_intensity_batch_audit.csv"
    pd.DataFrame(rows).to_csv(output_path, index=False)
    return output_path, rows


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("batch_root", type=Path)
    parser.add_argument("--exclude-target", action="append", default=[])
    args = parser.parse_args()
    output_path, rows = audit(args.batch_root, set(args.exclude_target))
    passed = sum(row["status"] == "pass" for row in rows)
    failed = sum(row["status"] == "fail" for row in rows)
    excluded = sum(row["status"] == "excluded" for row in rows)
    print(f"AUDIT={output_path}")
    print(f"PASS={passed} FAIL={failed} EXCLUDED={excluded}")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
