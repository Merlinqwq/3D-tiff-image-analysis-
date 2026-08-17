"""Read-only reconciliation of DAPI nuclei and legacy target-object CSV counts."""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
import tifffile
from openpyxl import load_workbook

from nucleus_intensity import filter_labels_by_area, split_touching_nuclei


def date_from_path(path: Path) -> str:
    for part in path.parts:
        match = re.match(r"^(\d{8})\s+MCF10A", part, re.IGNORECASE)
        if match:
            return match.group(1)
    return "unknown"


def group_name(date: str, ch2: str | None, ch3: str | None) -> str:
    targets = [value for value in (ch2, ch3) if value]
    if date == "20260717" and {value.casefold() for value in targets} == {"nop16", "pml"}:
        return "PML"
    return "+".join(targets)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("root", type=Path)
    parser.add_argument("--watershed-min-distance", type=int)
    parser.add_argument("--min-nucleus-radius", type=float, default=5.0)
    parser.add_argument("--min-area", type=int, default=5000)
    args = parser.parse_args()
    root = args.root.resolve()
    workbooks = sorted(
        path for path in root.rglob("nuclear_intensity_results.xlsx")
        if path.parent.name == "Intensity"
        and "ubf" not in path.as_posix().casefold()
        and "srrm1" not in path.as_posix().casefold()
    )

    nucleus_by_group = defaultdict(int)
    original_nucleus_by_group = defaultdict(int)
    image_rows: list[dict] = []
    missing_csv: list[str] = []
    csv_channels = Counter()
    legacy_object_totals = defaultdict(int)

    for workbook_path in workbooks:
        date = date_from_path(workbook_path)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook["Images"]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        index = {name: position for position, name in enumerate(headers)}
        workbook_images = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            group = group_name(date, row[index["channel_2_target"]], row[index["channel_3_target"]])
            key = (group, date, row[index["condition"]])
            original_count = int(row[index["nucleus_count"]])
            original_nucleus_by_group[key] += original_count
            if args.watershed_min_distance is not None:
                mask_path = workbook_path.parent / row[index["binary_mask_file"]]
                binary = tifffile.imread(mask_path) > 0
                checked_labels = split_touching_nuclei(
                        binary,
                        args.watershed_min_distance,
                        args.min_nucleus_radius,
                    )
                checked_labels = filter_labels_by_area(checked_labels, args.min_area)
                nucleus_count = int(checked_labels.max(initial=0))
            else:
                nucleus_count = original_count
            nucleus_by_group[key] += nucleus_count
            workbook_images.append(
                {
                    "image": row[index["image"]],
                    "group": group,
                    "date": date,
                    "condition": row[index["condition"]],
                    "nucleus_count": nucleus_count,
                }
            )
        workbook.close()

        new_folder = workbook_path.parent.parent
        csv_path = new_folder / "channel_analysis_results" / "combined_channel_object_analysis.csv"
        csv_counts: dict[str, int] = {}
        if csv_path.exists():
            legacy = pd.read_csv(csv_path)
            for channel, count in legacy.groupby("channel").size().items():
                csv_channels[str(channel)] += int(count)
            consistency = legacy.groupby(["image_name", "channel"])["object_count"].nunique()
            if (consistency > 1).any():
                raise ValueError(f"Inconsistent object_count within {csv_path}")
            per_image = legacy.groupby("image_name")["object_id"].count().to_dict()
            csv_counts = {str(name): int(value) for name, value in per_image.items()}
        else:
            missing_csv.append(str(new_folder))

        for record in workbook_images:
            legacy_count = csv_counts.get(record["image"])
            record["legacy_target_object_count"] = legacy_count
            if legacy_count is not None:
                legacy_object_totals[(record["group"], record["date"], record["condition"])] += legacy_count
            image_rows.append(record)

    print("DAPI_NUCLEI_BY_REPLICATE_TREATMENT")
    print("target_group\tdate\tcondition\toriginal_nuclei\tchecked_nuclei\tdeviation_from_25")
    for key in sorted(nucleus_by_group):
        count = nucleus_by_group[key]
        print(
            f"{key[0]}\t{key[1]}\t{key[2]}\t{original_nucleus_by_group[key]}\t"
            f"{count}\t{count - 25:+d}"
        )

    print("\nTARGET_TOTALS")
    target_totals = defaultdict(int)
    for (group, _, _), count in nucleus_by_group.items():
        target_totals[group] += count
    for group in sorted(target_totals):
        replicate_treatments = sum(1 for key in nucleus_by_group if key[0] == group)
        print(
            f"{group}\ttotal_nuclei={target_totals[group]}\t"
            f"replicate_treatments={replicate_treatments}\t"
            f"mean={target_totals[group] / replicate_treatments:.1f}"
        )

    paired = [row for row in image_rows if row["legacy_target_object_count"] is not None]
    if paired:
        nuclei = np.array([row["nucleus_count"] for row in paired], dtype=float)
        objects = np.array([row["legacy_target_object_count"] for row in paired], dtype=float)
        correlation = float(np.corrcoef(nuclei, objects)[0, 1]) if len(paired) > 1 else float("nan")
        exact = int(np.sum(nuclei == objects))
        print("\nLEGACY_CSV_COMPARISON")
        print(f"paired_images={len(paired)}")
        print(f"exact_count_matches={exact}")
        print(f"pearson_correlation={correlation:.3f}")
        print(f"median_dapi_nuclei_per_image={np.median(nuclei):.1f}")
        print(f"median_legacy_target_objects_per_image={np.median(objects):.1f}")
        print(f"legacy_channels={dict(csv_channels)}")
    print(f"csv_folders_found={len(workbooks) - len(missing_csv)}")
    print(f"csv_folders_missing={len(missing_csv)}")
    for path in missing_csv:
        print(f"MISSING_CSV\t{path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
