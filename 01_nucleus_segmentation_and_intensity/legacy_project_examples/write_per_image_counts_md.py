"""Generate a Markdown record of DAPI nucleus counts for every processed image."""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def acquisition_date(path: Path) -> str:
    for part in path.parts:
        match = re.match(r"^(\d{8})\s+MCF10A", part, re.IGNORECASE)
        if match:
            value = match.group(1)
            return f"{value[:4]}-{value[4:6]}-{value[6:]}"
    raise ValueError(f"Cannot infer acquisition date from {path}")


def target_group(date: str, channel_2: str | None, channel_3: str | None) -> str:
    targets = [value for value in (channel_2, channel_3) if value]
    if date == "2026-07-17" and {value.casefold() for value in targets} == {"nop16", "pml"}:
        return "PML"
    return " + ".join(targets)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("root", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    root = args.root.resolve()
    records: list[dict] = []

    workbooks = sorted(
        path for path in root.rglob("nuclear_intensity_results.xlsx")
        if path.parent.name == "Intensity"
        and "ubf" not in path.as_posix().casefold()
        and "srrm1" not in path.as_posix().casefold()
    )
    for workbook_path in workbooks:
        date = acquisition_date(workbook_path)
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        sheet = workbook["Images"]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        index = {name: position for position, name in enumerate(headers)}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            records.append(
                {
                    "target": target_group(
                        date,
                        row[index["channel_2_target"]],
                        row[index["channel_3_target"]],
                    ),
                    "date": date,
                    "condition": row[index["condition"]],
                    "image": row[index["image"]],
                    "nuclei": int(row[index["nucleus_count"]]),
                }
            )
        workbook.close()

    records.sort(key=lambda row: (row["target"].casefold(), row["date"], row["condition"], row["image"]))
    grouped: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for record in records:
        grouped[(record["target"], record["date"], record["condition"])].append(record)

    lines = [
        "# DAPI nucleus count per image",
        "",
        "Generated from each `Intensity/nuclear_intensity_results.xlsx` workbook.",
        "Counts are labeled DAPI nuclear ROIs after sum projection, Otsu thresholding, hole filling, watershed separation, and the 5,000-pixel minimum-area filter.",
        "UBF and SRRM1 are excluded. For 2026-07-17 PML/NOP16 images, the group is recorded as PML as requested.",
        "",
        f"- Processed images: {len(records)}",
        f"- Total nuclei: {sum(row['nuclei'] for row in records)}",
        "",
    ]
    for (target, date, condition), group_records in grouped.items():
        subtotal = sum(row["nuclei"] for row in group_records)
        lines.extend(
            [
                f"## {target} — {date} — {condition}",
                "",
                "| Image | Nuclei |",
                "|---|---:|",
                *[f"| {row['image']} | {row['nuclei']} |" for row in group_records],
                f"| **Replicate-treatment subtotal** | **{subtotal}** |",
                "",
            ]
        )

    args.output.write_text("\n".join(lines), encoding="utf-8")
    print(f"OUTPUT={args.output.resolve()}")
    print(f"IMAGES={len(records)}")
    print(f"NUCLEI={sum(row['nuclei'] for row in records)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
