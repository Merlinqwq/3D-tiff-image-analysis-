"""Generic per-object-ROI 3D colocalization analysis for two user-defined channels.

The script reuses existing 2D object label masks and two existing 3D binary
condensate masks. It measures background-corrected Pearson correlation,
binary-mask overlap, object overlap, and anisotropy-corrected centroid distance.
Source TIFFs and prior masks are read-only.
"""

from __future__ import annotations

import argparse
import importlib.util
import math
from pathlib import Path

import numpy as np
import pandas as pd
import tifffile
from scipy.spatial.distance import cdist
from skimage.measure import label, regionprops


ROOT = Path(__file__).resolve().parent
SPECIALIZED_PATH = ROOT / "background_corrected_condensate_thresholding.py"
spec = importlib.util.spec_from_file_location("background_corrected", SPECIALIZED_PATH)
background_corrected = importlib.util.module_from_spec(spec)
spec.loader.exec_module(background_corrected)
base = background_corrected.base

OUTPUT_FOLDER = "colocalization_analysis"
MERGED_FOLDER = "two_channel_colocalization_results"
MERGED_WORKBOOK = "Merged_two_channel_3D_colocalization.xlsx"


def pearson_r(a: np.ndarray, b: np.ndarray) -> float:
    x = np.asarray(a, dtype=np.float64).ravel()
    y = np.asarray(b, dtype=np.float64).ravel()
    valid = np.isfinite(x) & np.isfinite(y)
    x, y = x[valid], y[valid]
    if x.size < 2 or np.ptp(x) == 0 or np.ptp(y) == 0:
        return math.nan
    return float(np.corrcoef(x, y)[0, 1])


def safe_fraction(numerator: int, denominator: int) -> float:
    return float(numerator / denominator) if denominator else math.nan


def mask_metrics(mask_a: np.ndarray, mask_b: np.ndarray) -> dict[str, float | int]:
    count_a = int(np.count_nonzero(mask_a))
    count_b = int(np.count_nonzero(mask_b))
    intersection = int(np.count_nonzero(mask_a & mask_b))
    union = int(np.count_nonzero(mask_a | mask_b))
    return {
        "target_a_condensate_voxels": count_a,
        "target_b_condensate_voxels": count_b,
        "overlap_voxels": intersection,
        "union_voxels": union,
        "dice_coefficient": safe_fraction(2 * intersection, count_a + count_b),
        "jaccard_index": safe_fraction(intersection, union),
        "target_a_volume_fraction_overlapping_target_b": safe_fraction(intersection, count_a),
        "target_b_volume_fraction_overlapping_target_a": safe_fraction(intersection, count_b),
    }


def object_metrics(
    mask_a: np.ndarray, mask_b: np.ndarray, z_step_um: float, xy_step_um: float
) -> dict[str, float | int]:
    labels_a = label(mask_a, connectivity=1)
    labels_b = label(mask_b, connectivity=1)
    props_a = regionprops(labels_a)
    props_b = regionprops(labels_b)
    overlap_a = sum(bool(np.any(mask_b[labels_a == item.label])) for item in props_a)
    overlap_b = sum(bool(np.any(mask_a[labels_b == item.label])) for item in props_b)
    centroids_a = np.asarray([item.centroid for item in props_a], dtype=float)
    centroids_b = np.asarray([item.centroid for item in props_b], dtype=float)
    a_to_b = b_to_a = math.nan
    if len(centroids_a) and len(centroids_b):
        scale = np.asarray([z_step_um, xy_step_um, xy_step_um])
        distances = cdist(centroids_a * scale, centroids_b * scale)
        a_to_b = float(np.mean(np.min(distances, axis=1)))
        b_to_a = float(np.mean(np.min(distances, axis=0)))
    return {
        "target_a_object_count": len(props_a),
        "target_b_object_count": len(props_b),
        "target_a_objects_overlapping_target_b": overlap_a,
        "target_b_objects_overlapping_target_a": overlap_b,
        "target_a_object_fraction_overlapping_target_b": safe_fraction(overlap_a, len(props_a)),
        "target_b_object_fraction_overlapping_target_a": safe_fraction(overlap_b, len(props_b)),
        "mean_target_a_to_nearest_target_b_centroid_um": a_to_b,
        "mean_target_b_to_nearest_target_a_centroid_um": b_to_a,
    }


def focus_mask_path(folder: Path, image: Path, target: str, channel: int) -> Path:
    path = (
        folder / base.RESULT_FOLDER_NAME / "Masks" /
        f"{image.stem}_{target}_C{channel}_whole_image_foci_binary_mask.tif"
    )
    if not path.exists():
        raise FileNotFoundError(f"Missing {target} mask: {path}")
    return path


def selected_folders(
    batch_root: Path, target_a: str, channel_a: int,
    target_b: str, channel_b: int, folder_name: str | None,
) -> list[Path]:
    selected: list[Path] = []
    for folder in base.find_analysis_folders(batch_root, folder_name):
        images = base._tiff_files(folder)
        if not images:
            continue
        try:
            focus_mask_path(folder, images[0], target_a, channel_a)
            focus_mask_path(folder, images[0], target_b, channel_b)
        except FileNotFoundError:
            continue
        else:
            selected.append(folder)
    return selected


def analyze_image(
    folder: Path, image: Path, replicate: str,
    target_a: str, channel_a: int, target_b: str, channel_b: int,
    xy_pixel_size_um: float, z_step_um: float,
    background_window_a: int, background_window_b: int,
) -> list[dict]:
    roi_labels = tifffile.imread(base._label_mask_for_image(folder, image)).astype(np.int32)
    binary_a = tifffile.imread(focus_mask_path(folder, image, target_a, channel_a)) > 0
    binary_b = tifffile.imread(focus_mask_path(folder, image, target_b, channel_b)) > 0
    with tifffile.TiffFile(image) as tif:
        series = tif.series[0]
        zcyx = base._normalize_to_zcyx(series.asarray(), series.axes)
    if max(channel_a, channel_b) > zcyx.shape[1]:
        raise ValueError(
            f"Requested channel {max(channel_a, channel_b)}, but {image.name} has {zcyx.shape[1]} channels"
        )
    if binary_a.shape != binary_b.shape or binary_a.shape != zcyx[:, 0].shape:
        raise ValueError(
            f"Shape mismatch image={zcyx.shape}, target_a={binary_a.shape}, target_b={binary_b.shape}"
        )
    if roi_labels.shape != zcyx.shape[-2:]:
        raise ValueError(f"ROI shape {roi_labels.shape} != image YX {zcyx.shape[-2:]}")

    raw_a = zcyx[:, channel_a - 1].astype(np.float32, copy=False)
    raw_b = zcyx[:, channel_b - 1].astype(np.float32, copy=False)
    residual_a, _ = background_corrected.background_residual(
        raw_a, target_a, background_window_a
    )
    residual_b, _ = background_corrected.background_residual(
        raw_b, target_b, background_window_b
    )
    voxel_volume = xy_pixel_size_um ** 2 * z_step_um
    rows: list[dict] = []
    for roi in regionprops(roi_labels):
        roi_2d = roi_labels == roi.label
        roi_3d = np.broadcast_to(roi_2d, binary_a.shape)
        mask_a = binary_a & roi_3d
        mask_b = binary_b & roi_3d
        union = mask_a | mask_b
        overlap = mask_metrics(mask_a, mask_b)
        objects = object_metrics(mask_a, mask_b, z_step_um, xy_pixel_size_um)
        rows.append({
            "acquisition_date": base._date_from_path(folder),
            "condition": base._condition_from_folder(folder),
            "replicate": replicate,
            "source_folder": str(folder),
            "image_name": image.name,
            "roi_id": int(roi.label),
            "roi_area_px": int(roi.area),
            "roi_voxels_analyzed": int(np.count_nonzero(roi_3d)),
            "target_a": target_a,
            "target_b": target_b,
            "channel_a": channel_a,
            "channel_b": channel_b,
            "pearson_r_background_corrected_whole_roi": pearson_r(
                residual_a[roi_3d], residual_b[roi_3d]
            ),
            "pearson_r_raw_whole_roi_audit": pearson_r(raw_a[roi_3d], raw_b[roi_3d]),
            "pearson_r_background_corrected_mask_union": (
                pearson_r(residual_a[union], residual_b[union]) if np.any(union) else math.nan
            ),
            **overlap,
            "target_a_condensate_volume_um3": overlap["target_a_condensate_voxels"] * voxel_volume,
            "target_b_condensate_volume_um3": overlap["target_b_condensate_voxels"] * voxel_volume,
            "overlap_volume_um3": overlap["overlap_voxels"] * voxel_volume,
            **objects,
            "xy_pixel_size_um": xy_pixel_size_um,
            "z_step_um": z_step_um,
            "background_window_a_px": background_window_a,
            "background_window_b_px": background_window_b,
            "gaussian_sigma_px": background_corrected.GAUSSIAN_SIGMA,
        })
    return rows


def image_summary(rows: pd.DataFrame) -> pd.DataFrame:
    metadata = [
        "acquisition_date", "condition", "replicate", "source_folder",
        "image_name", "target_a", "target_b",
    ]
    metrics = [
        "pearson_r_background_corrected_whole_roi",
        "pearson_r_background_corrected_mask_union",
        "dice_coefficient", "jaccard_index",
        "target_a_volume_fraction_overlapping_target_b",
        "target_b_volume_fraction_overlapping_target_a",
        "target_a_object_fraction_overlapping_target_b",
        "target_b_object_fraction_overlapping_target_a",
        "mean_target_a_to_nearest_target_b_centroid_um",
        "mean_target_b_to_nearest_target_a_centroid_um",
    ]
    grouped = rows.groupby(metadata, dropna=False)
    output = grouped.size().rename("roi_count").reset_index()
    for metric in metrics:
        stats = grouped[metric].agg(["mean", "sem"]).reset_index().rename(
            columns={"mean": f"{metric}_mean", "sem": f"{metric}_sem"}
        )
        output = output.merge(stats, on=metadata, how="left")
    return output


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill, cell.font = fill, font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in sheet.columns:
            values = [str(cell.value or "") for cell in column[:200]]
            width = min(max(max(map(len, values), default=0) + 2, 12), 55)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = width
    workbook.save(path)


def run(
    batch_root: Path, target_a: str, channel_a: int,
    target_b: str, channel_b: int, xy_pixel_size_um: float, z_step_um: float,
    background_window_a: int, background_window_b: int,
    folder_name: str | None = None,
) -> Path:
    batch_root = Path(batch_root).resolve()
    folders = selected_folders(
        batch_root, target_a, channel_a, target_b, channel_b, folder_name
    )
    if not folders:
        raise FileNotFoundError("No selected folder contains both requested masks")
    replicate_map = base.build_replicate_map(base.find_analysis_folders(batch_root, folder_name))
    all_rows: list[dict] = []
    errors: list[dict] = []
    for folder_index, folder in enumerate(folders, start=1):
        date = base._date_from_path(folder)
        replicate = replicate_map[(base._replicate_group_key(folder), date)]
        folder_rows: list[dict] = []
        images = base._tiff_files(folder)
        print(f"DATASET [{folder_index}/{len(folders)}] {folder}")
        for image_index, image in enumerate(images, start=1):
            print(f"  [{image_index}/{len(images)}] {image.name}")
            try:
                folder_rows.extend(analyze_image(
                    folder, image, replicate, target_a, channel_a, target_b, channel_b,
                    xy_pixel_size_um, z_step_um, background_window_a, background_window_b,
                ))
            except Exception as exc:
                errors.append({
                    "source_folder": str(folder), "image_name": image.name,
                    "error": f"{type(exc).__name__}: {exc}",
                })
        output = folder / OUTPUT_FOLDER
        output.mkdir(exist_ok=True)
        folder_frame = pd.DataFrame(folder_rows)
        folder_frame.to_csv(output / "per_roi_colocalization.csv", index=False)
        if not folder_frame.empty:
            image_summary(folder_frame).to_csv(output / "per_image_summary.csv", index=False)
        pd.DataFrame([{
            "target_a": target_a, "channel_a": channel_a,
            "target_b": target_b, "channel_b": channel_b,
            "xy_pixel_size_um": xy_pixel_size_um, "z_step_um": z_step_um,
            "background_window_a_px": background_window_a,
            "background_window_b_px": background_window_b,
            "gaussian_sigma_px": background_corrected.GAUSSIAN_SIGMA,
            "existing_masks_modified": False,
        }]).to_csv(output / "method_parameters.csv", index=False)
        all_rows.extend(folder_rows)

    merged = pd.DataFrame(all_rows)
    summary = image_summary(merged) if not merged.empty else pd.DataFrame()
    results_dir = batch_root / MERGED_FOLDER
    results_dir.mkdir(exist_ok=True)
    merged.to_csv(results_dir / "per_roi_colocalization.csv", index=False)
    summary.to_csv(results_dir / "per_image_summary.csv", index=False)
    error_frame = pd.DataFrame(errors, columns=["source_folder", "image_name", "error"])
    error_frame.to_csv(results_dir / "errors.csv", index=False)
    method = pd.DataFrame([
        {"parameter": "target_a", "value": target_a},
        {"parameter": "channel_a", "value": channel_a},
        {"parameter": "target_b", "value": target_b},
        {"parameter": "channel_b", "value": channel_b},
        {"parameter": "xy_pixel_size_um", "value": xy_pixel_size_um},
        {"parameter": "z_step_um", "value": z_step_um},
        {"parameter": "background_window_a_px", "value": background_window_a},
        {"parameter": "background_window_b_px", "value": background_window_b},
        {"parameter": "existing_masks_modified", "value": False},
    ])
    workbook = batch_root / MERGED_WORKBOOK
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="Per ROI", index=False)
        summary.to_excel(writer, sheet_name="Per image summary", index=False)
        method.to_excel(writer, sheet_name="Method", index=False)
        error_frame.to_excel(writer, sheet_name="Errors", index=False)
    format_workbook(workbook)
    print(f"COMPLETE rois={len(merged)} errors={len(errors)}")
    print(f"WORKBOOK {workbook}")
    return workbook


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--batch-root", type=Path,
        help="Selected TIFF folder or root containing TIFF folders. Opens a picker if omitted.",
    )
    parser.add_argument("--folder-name", help="Optional exact folder-name restriction")
    parser.add_argument(
        "--parameter-table", type=Path,
        default=base.PARAMETER_TABLE_RELATIVE_PATH,
        help="Parameter CSV path relative to each selected TIFF folder.",
    )
    parser.add_argument("--target-a", default="TargetA")
    parser.add_argument("--channel-a", type=int, default=2)
    parser.add_argument("--target-b", default="TargetB")
    parser.add_argument("--channel-b", type=int, default=3)
    parser.add_argument("--xy-pixel-size", type=float, help="XY pixel size in micrometers")
    parser.add_argument("--z-step", type=float, help="Z spacing in micrometers")
    parser.add_argument("--background-window-a", type=int, default=31)
    parser.add_argument("--background-window-b", type=int, default=31)
    args = parser.parse_args()
    using_gui = args.batch_root is None
    batch_root = args.batch_root or base.choose_folder_gui()
    if batch_root is None:
        return 0
    xy_pixel_size = args.xy_pixel_size
    z_step = args.z_step
    if using_gui:
        xy_pixel_size = xy_pixel_size or base.choose_positive_float_gui(
            "XY calibration", "XY pixel size in micrometers:"
        )
        z_step = z_step or base.choose_positive_float_gui(
            "Z calibration", "Z step in micrometers:"
        )
        if xy_pixel_size is None or z_step is None:
            return 0
    if xy_pixel_size is None or z_step is None:
        raise SystemExit("Provide --xy-pixel-size and --z-step in command-line mode")
    if args.channel_a < 1 or args.channel_b < 1 or args.channel_a == args.channel_b:
        raise SystemExit("Channels must be distinct positive one-based integers")
    if xy_pixel_size <= 0 or z_step <= 0:
        raise SystemExit("Pixel size and Z step must be positive")
    for window in (args.background_window_a, args.background_window_b):
        if window < 3 or window % 2 == 0:
            raise SystemExit("Background windows must be odd integers >= 3")
    if args.parameter_table.is_absolute():
        raise SystemExit("--parameter-table must be relative to each selected TIFF folder")
    base.PARAMETER_TABLE_RELATIVE_PATH = args.parameter_table
    run(
        batch_root, args.target_a, args.channel_a, args.target_b, args.channel_b,
        xy_pixel_size, z_step, args.background_window_a,
        args.background_window_b, args.folder_name,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
